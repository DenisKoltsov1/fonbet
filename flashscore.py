from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from fake_useragent import UserAgent
from selenium.webdriver.common.keys import Keys
import undetected_chromedriver 
import os
options = webdriver.ChromeOptions()
driver=undetected_chromedriver.Chrome()
import pandas as pd
import glob


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string



def parser(driver,id,name,price):
      
           
    time.sleep(5)
    # переход  на сайт  ozon
    html=driver.get("https://www.ozon.ru/")
    time.sleep(5)
    #ищем  поле поиска
    kcal = driver.find_element(By.XPATH,"//input[contains(@class,'tsBodyL')]")
    time.sleep(5)
    #вставляем  id
    kcal.send_keys(id)
    time.sleep(5)
    #лупа для продолжения поиска
    lupa = driver.find_element(By.XPATH,'//button[@class="x5-a1 x5-f0"]')
    time.sleep(5)
    lupa.click()

    time.sleep(5)

    # клик по ссылке  продукта 
    item=driver.find_element(By.XPATH,"//span[contains(@class,'tsBodyL')]")
    item.click()
    time.sleep(5)
    # ищем имя  товара
    try: 
        name=driver.find_element(By.XPATH,"//h1[contains(@class,'rn')]").text
        print(name) 
    except:
        name=''
        print('элемент name(имя) не найден')     
    
    #ищем цену  товара
    try:
        price=driver.find_element(By.XPATH,"//span[contains(@class,'p1n np2')]").text.replace("₽", "")
        print(price)
    except:
        price=''
        print('элемент price(цена) не найден')
    
    #ищем цену  до скидки товара
    try:
        priceSkidka=driver.find_element(By.XPATH,"//span[contains(@class,'pn2')]").text.replace("₽", "")   #pn2
        print(priceSkidka)      
    except:
        priceSkidka=''
        print('элемент priceSkidka(цена со скидкой) не найден')   
          
    #ищем  nds
    nds='13%'
    print(nds)
 

    #KomercialType
    try:
        KomercialType=driver.find_element(By.XPATH,"//span[text()='Тип']/following::dd[1]").text
        print(KomercialType.text)
    except:
        KomercialType=''
        print('элемент KomercialType не найден')


    #SerialNumber
    #SerialNumber=driver.find_element(By.XPATH,"//span[text()='Тип']/following::dd[1]")
    #print(SerialNumber.text)
    #ищем  id 
    Ozonid=id
    print(Ozonid)
    #ищем тип  товара 
    #


    #ссылка на главное фото 
    #mainFoto=driver.find_element(By.XPATH,"//input[contains(@class,'tsBodyL')]")")
    #бренд   товара
    brendItem=driver.find_element(By.XPATH,"//span[text()='Бренд']/following::dd[1]")
    print(brendItem.text)  
    time.sleep(45)

    #тип  товара
    #/td/span[text()='Данные2']/following::td[1]/span
    typeItem=driver.find_element(By.XPATH,"//span[text()='Тип']/following::dd[1]")
    print(typeItem.text)

    #объем товара
    obiem=driver.find_element(By.XPATH,"//span[text()='Объем, мл']/following::dd[1]")
    print(obiem.text)
    
    #едениц товара
    countItem=1
    print(countItem)  

    #пол
    femal=driver.find_element(By.XPATH,"//span[text()='Целевая аудитория']/following::dd[1]")
    print(femal.text)  

    #класс опасности
    classop="не опасен"
    print(classop)
    

    #страна изготовитель
    
    ManufacturerСountry=driver.find_element(By.XPATH,"//span[text()='Страна-изготовитель']/following::dd[1]")
    print(ManufacturerСountry.text)
       
    return id,name,price,priceSkidka,nds,KomercialType
    #return 1
 



def OpenPixx(file,id,c,r,name,price,priceSkidka,nds,KomercialType):
    hel='hello'
    wb = load_workbook(file)
    # get_sheet_names() - выводит список с названием листов,
    ws = wb['Шаблон для поставщика']
    #ws['B4']=id
    
    for c in range(3,40):

        ws.cell(row=r, column=c).value = id
        ws.cell(row=r, column=c).value = name
        ws.cell(row=r, column=c).value = price
        ws.cell(row=r, column=c).value = priceSkidka
        ws.cell(row=r, column=c).value = nds
        ws.cell(row=r, column=c).value = id
        ws.cell(row=r, column=c).value = KomercialType
        ws.cell(row=r, column=c).value = priceSkidka
        ws.cell(row=r, column=c).value = priceSkidka
        ws.cell(row=r, column=c).value = priceSkidka
   
    wb.save(file)
    return 1
name=''
price=''
priceSkidka=''
nds=''
KomercialType=''
id=215973958    
parser(driver,id,name,price)
file ="C:\\ozon1\\Парфюмерия.xlsx"
print(name,price,priceSkidka,nds,KomercialType)
#openExel(file)
r=6
c=3
OpenPixx(file,id,c,r,name,price,priceSkidka,nds,KomercialType)

